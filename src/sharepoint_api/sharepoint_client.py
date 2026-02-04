import os
from typing import Union

import msal
import requests
import json
import pandas as pd
import io
import toml
import streamlit as st
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta, timezone




class SharePointClient:
    def __init__(self, site:str, dry_run:bool=None):
        self.site = site
        self.is_local = self._is_local()
        self.dry_run = dry_run
        if dry_run is None:
            self.dry_run = self.is_local
        self._load_config()
        self._authenticate()
        self._get_site_and_drive_ids()

    @staticmethod
    def _is_local():
        # Example heuristic: check environment variable or file existence
        env_path = os.getenv("SECRETS_PATH")
        if env_path and os.path.exists(env_path):
            return env_path
        return False

    def _load_config(self):
        if self.is_local:
            config = toml.load(self.is_local)
            config_azure = config.get("azure", {})
            config_site = config.get(self.site, {})
        else:
            config_azure = st.secrets.get("azure", {})
            config_site = st.secrets.get("azure", {})

        self.tenant_id = config_azure["tenant_id"]
        self.client_id = config_azure["client_id"]
        self.client_secret = config_azure["client_secret"]
        self.site_domain = config_site["site_domain"]
        self.site_name = config_site["site_name"]

    def _authenticate(self):
        authority = f"https://login.microsoftonline.com/{self.tenant_id}"
        app = msal.ConfidentialClientApplication(
            self.client_id, authority=authority,
            client_credential=self.client_secret
        )
        scopes = ["https://graph.microsoft.com/.default"]
        result = app.acquire_token_for_client(scopes=scopes)
        self.access_token = result['access_token']
        self.token_expiry = datetime.now(timezone.utc) + timedelta(seconds=result["expires_in"])
        self.headers = {"Authorization": f"Bearer {self.access_token}"}

    def _get_site_and_drive_ids(self):
        site_url = f"https://graph.microsoft.com/v1.0/sites/{self.site_domain}:/sites/{self.site_name}"
        site_resp = requests.get(site_url, headers=self.headers)
        self.site_id = site_resp.json()["id"]

        drive_url = f"https://graph.microsoft.com/v1.0/sites/{self.site_id}/drive"
        drive_resp = requests.get(drive_url, headers=self.headers)
        self.drive_id = drive_resp.json()["id"]

    def read_excel(self, file_path: str, sheet_name: str = 0) -> pd.DataFrame:
        file_url = f"https://graph.microsoft.com/v1.0/drives/{self.drive_id}/root:/{file_path}:/content"
        response = requests.get(file_url, headers=self.headers)
        response.raise_for_status()  # Raise an error for bad responses
        df = pd.read_excel(io.BytesIO(response.content), sheet_name=sheet_name, engine="openpyxl")
        return df

    def save_excel(self, df: pd.DataFrame, file_path: str, sheet_name: str = "Sheet1", header: bool = True):
        if self.dry_run:
            return
        buffer = io.BytesIO()
        df.to_excel(buffer, index=False, na_rep="", header=header, sheet_name=sheet_name, engine="openpyxl")
        buffer.seek(0)

        upload_url = f"https://graph.microsoft.com/v1.0/drives/{self.drive_id}/root:/{file_path}:/content"
        response = requests.put(upload_url, headers=self.headers, data=buffer.read())
        if response.status_code not in [200, 201]:
            raise Exception(f"Failed to update file: {file_path}")
        return response.status_code in [200, 201]


    def read_csv(self, file_path: str, sep: str = ',', encoding: str = 'utf-8') -> pd.DataFrame:
        """
        Reads a CSV file from SharePoint into a pandas DataFrame.

        :param file_path: Path to the CSV file in SharePoint (e.g., 'Logs/log_summary.csv').
        :param sep: Field delimiter (default ','; use '\t' for tab-delimited files).
        :param encoding: File encoding (default 'utf-8').
        :return: pandas DataFrame.
        """
        file_url = f"https://graph.microsoft.com/v1.0/drives/{self.drive_id}/root:/{file_path}:/content"
        response = requests.get(file_url, headers=self.headers)
        response.raise_for_status()  # Raise exception for HTTP errors

        buffer = io.BytesIO(response.content)
        try:
            df = pd.read_csv(buffer, sep=sep, encoding=encoding)
        except UnicodeDecodeError:
            buffer.seek(0)
            df = pd.read_csv(buffer, sep=sep, encoding='latin1')
        return df

    def save_csv(self, df: pd.DataFrame, file_path: str, sep: str = ','):
        """
        Saves a DataFrame as a CSV file to SharePoint.

        :param df: pandas DataFrame to save.
        :param file_path: Path in SharePoint to save the CSV file.
        :param sep: Field delimiter (default ','; use '\t' for tab-delimited .txt).
        :return: True if upload succeeded, False otherwise.
        """
        if self.dry_run:
            return
        buffer = io.StringIO()
        df.to_csv(buffer, index=False, na_rep="", sep=sep)
        buffer.seek(0)
        content_type = "text/plain" if file_path.endswith(".txt") else "text/csv"

        upload_url = f"https://graph.microsoft.com/v1.0/drives/{self.drive_id}/root:/{file_path}:/content"
        response = requests.put(
            upload_url,
            headers={**self.headers, "Content-Type": content_type},
            data=buffer.getvalue()
        )
        if response.status_code not in [200, 201]:
            raise Exception(f"Failed to update file: {file_path}")
        return response.status_code in [200, 201]

    def read_json(self, file_path: str) -> dict:
        """
        Reads a JSON file from SharePoint and returns it as a Python dictionary.

        :param file_path: Path to the JSON file in SharePoint.
        :return: Parsed JSON content as a Python dict.
        """
        file_url = f"https://graph.microsoft.com/v1.0/drives/{self.drive_id}/root:/{file_path}:/content"
        response = requests.get(file_url, headers=self.headers)
        response.raise_for_status()
        return response.json()

    def save_json(self, data: dict, file_path: str, save_local: bool = False):
        """
        Saves a Python dictionary as a JSON file to SharePoint.

        :param data: Python dictionary to save.
        :param file_path: Path in SharePoint to save the JSON file.
        :param save_local: If True, save the local file. Default is False.
        :return: True if upload succeeded, False otherwise.
        """
        if self.dry_run and not save_local:
            return True
        upload_url = f"https://graph.microsoft.com/v1.0/drives/{self.drive_id}/root:/{file_path}:/content"
        response = requests.put(
            upload_url,
            headers={**self.headers, "Content-Type": "application/json"},
            data=json.dumps(data)
        )
        return response.status_code in [200, 201]

    def save_multiple_dfs_to_excel(self, dfs: list, sheet_names: list, file_path: str,
                                   auto_adjust_columns: bool = False):
        """
        Saves multiple DataFrames to different sheets in a single Excel file on SharePoint, with optional column width adjustment.

        :param dfs: List of pandas DataFrames.
        :param sheet_names: List of sheet names.
        :param file_path: Path in SharePoint to save the Excel file.
        :param auto_adjust_columns: Whether to auto-adjust column widths for readability.
        :return: True if upload succeeded, False otherwise.
        """
        if self.dry_run:
            return
        if len(dfs) != len(sheet_names):
            raise ValueError("Number of DataFrames and sheet names must be equal.")

        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            for df, sheet_name in zip(dfs, sheet_names):
                if isinstance(df, pd.DataFrame):
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            writer.book.save(buffer)  # Save the writer content to buffer

        if auto_adjust_columns:
            buffer.seek(0)
            wb = load_workbook(buffer)
            self.autoadjust_column_widths(wb)
            buffer = io.BytesIO()
            wb.save(buffer)

            # Save adjusted workbook to a new buffer
            buffer = io.BytesIO()
            wb.save(buffer)

        buffer.seek(0)
        upload_url = f"https://graph.microsoft.com/v1.0/drives/{self.drive_id}/root:/{file_path}:/content"
        response = requests.put(upload_url, headers=self.headers, data=buffer.read())
        return response.status_code in [200, 201]



    @staticmethod
    def autoadjust_column_widths(workbook: Workbook):
        """
        Auto-adjusts column widths for all sheets in the given openpyxl Workbook.

        :param workbook: openpyxl Workbook object.
        """
        for ws in workbook.worksheets:
            for col_idx, column_cells in enumerate(ws.columns, start=1):
                max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
                adjusted_width = max_length + 2  # Add padding
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = adjusted_width

    def create_folder_path(self, folder_path: str):
        """
        Creates a folder path recursively in SharePoint.
        :param folder_path: Folder path relative to SharePoint root (e.g., 'OC/Customer/2025/07/1234')
        :return: True if created successfully or already exists.
        """
        if self.dry_run:
            return
        parts_lst = folder_path.strip("/").split("/")
        parent_path = parts_lst[0]
        parts = parts_lst[1:]

        for part in parts:
            current_path = f"{parent_path}/{part}" if parent_path else part
            url = f"https://graph.microsoft.com/v1.0/drives/{self.drive_id}/root:/{current_path}"
            # Check if folder exists
            check_resp = requests.get(url, headers=self.headers)
            if check_resp.status_code == 404:
                # Folder doesn't exist, create it
                create_url = f"https://graph.microsoft.com/v1.0/drives/{self.drive_id}/root:/{parent_path}:/children" if parent_path else f"https://graph.microsoft.com/v1.0/drives/{self.drive_id}/root/children"
                body = {
                    "name": part,
                    "folder": {},
                    "@microsoft.graph.conflictBehavior": "replace"
                }
                create_resp = requests.post(create_url, headers={**self.headers, "Content-Type": "application/json"},
                                            json=body)
                if create_resp.status_code not in [200, 201]:
                    raise Exception(f"Failed to create folder '{current_path}': {create_resp.text}")
            elif check_resp.status_code not in [200, 201]:
                raise Exception(f"Failed to check folder '{current_path}': {check_resp.text}")
            parent_path = current_path  # Move to next subfolder
        return True

    def rename_folder(self, folder_path: str, new_name: str):
        """
        Renames a folder in SharePoint.
        :param folder_path: Folder path relative to SharePoint root (e.g., 'OC/Customer/2025/07/1234')
        :param new_name: New name for the folder
        :return: True if renamed successfully.
        """
        if self.dry_run:
            return

        folder_path = folder_path.strip("/")
        url = f"https://graph.microsoft.com/v1.0/drives/{self.drive_id}/root:/{folder_path}"

        # Check if folder exists
        check_resp = requests.get(url, headers=self.headers)
        if check_resp.status_code == 404:
            raise Exception(f"Folder '{folder_path}' does not exist.")
        elif check_resp.status_code not in [200, 201]:
            raise Exception(f"Failed to check folder '{folder_path}': {check_resp.text}")

        # Rename folder
        patch_resp = requests.patch(
            url,
            headers={**self.headers, "Content-Type": "application/json"},
            json={"name": new_name}
        )

        if patch_resp.status_code not in [200, 201]:
            raise Exception(f"Failed to rename folder '{folder_path}' to '{new_name}': {patch_resp.text}")

        return True

    @staticmethod
    def _format_delivery_note(wb: Workbook):
        ws = wb['Sheet1']
        col_widths = [22, 20, 50, 15, 15, 15, 20]
        for col_letter, width in zip(list('ABCDEFG'), col_widths):
            ws.column_dimensions[col_letter].width = width

        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5

        for row in range(14, ws.max_row + 1):
            cell = ws[f"G{row}"]
            cell.number_format = '#,##0.00'

    def save_delivery_note_excel(self, df: pd.DataFrame, file_path: str, sheet_name: str = "Sheet1"):
        """
        Saves a delivery note Excel file with special formatting.
        """
        if self.dry_run:
            return
        buffer = io.BytesIO()
        df.to_excel(buffer, index=False, header=False, sheet_name=sheet_name, engine="openpyxl")
        buffer.seek(0)

        wb = load_workbook(buffer)
        self._format_delivery_note(wb)

        final_buffer = io.BytesIO()
        wb.save(final_buffer)
        final_buffer.seek(0)

        upload_url = f"https://graph.microsoft.com/v1.0/drives/{self.drive_id}/root:/{file_path}:/content"
        response = requests.put(upload_url, headers=self.headers, data=final_buffer.read())
        return response.status_code in [200, 201]

    def is_excel_file_locked(self, file_path: str) -> Union[bool, None]:
        """
        Check if the Excel file at file_path is currently locked (has an active workbook session).
        Returns:
            True if locked,
            False if not locked,
            None if not an Excel file or unknown error.
        """
        url = f"https://graph.microsoft.com/v1.0/sites/{self.site_id}/drive/root:/{file_path}:/workbook/createSession"
        headers = self.headers.copy()
        headers["Content-Type"] = "application/json"
        body = {"persistChanges": True}

        response = requests.post(url, headers=headers, json=body)

        if response.status_code == 201:
            # Success: workbook not locked; close session to avoid leaving it open
            session_id = response.json().get("id")
            close_url = f"https://graph.microsoft.com/v1.0/sites/{self.site_id}/drive/root:/{file_path}:/workbook/closeSession"
            close_headers = headers.copy()
            close_headers["workbook-session-id"] = session_id
            requests.post(close_url, headers=close_headers)
            return False

        elif response.status_code == 409:
            # 409 Conflict: workbook is currently locked (session already exists)
            return True

        elif response.status_code == 415:
            # 415 Unsupported Media Type: likely not an Excel file
            return None

        else:
            # Other unexpected error
            print(f"Error checking lock: {response.status_code} - {response.text}")
            return None

    @staticmethod
    def format_columns_no_scientific_notation(ws, columns):
        for col_idx in columns:
            col_letter = get_column_letter(col_idx)
            for cell in ws[col_letter]:
                if cell.row == 1:
                    continue  # Skip header
                cell.number_format = '0'  # Integer format (no decimals, no scientific)


    def list_files_in_folder(self, folder_path):
        """
        folder_path: path relative to root of the document library
        returns: list of dicts with file info
        """
        url = f"https://graph.microsoft.com/v1.0/drives/{self.drive_id}/root:/{folder_path}:/children"
        headers = self.headers.copy()
        r = requests.get(url, headers=headers)
        r.raise_for_status()
        return r.json().get("value", [])
